"""Lead + memory endpoints.

  GET  /leads                     list leads (optional ?status=, paging)
  POST /leads                     create/seed a lead
  GET  /leads/{id}                full lead view
  GET  /leads/{id}/memory         memory bundle for the brain (LEAD PROFILE slot)
"""

from __future__ import annotations

import io
import uuid

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from pydantic import BaseModel

from business.logging import get_logger
from business.models import FunnelStage, Lead, LeadStatus, is_lead_for_stage
from business.schemas import CreateLeadRequest, LeadView, MemoryBundle
from business.store import get_store

log = get_logger(__name__)
router = APIRouter(prefix="/leads", tags=["leads"])


def _to_view(lead: Lead) -> LeadView:
    return LeadView(
        lead_id=lead.lead_id,
        full_name=lead.full_name,
        email=lead.email,
        phone_e164=lead.phone_e164,
        source=lead.source,
        language_preference=lead.language_preference,
        course_interest=lead.course_interest,
        intake_year=lead.intake_year,
        city=lead.city,
        consent_call=lead.consent_call,
        consent_whatsapp=lead.consent_whatsapp,
        status=lead.status,
        funnel_stage=lead.funnel_stage,
        lead_priority=lead.lead_priority,
        is_lead=lead.is_lead,
        facts=lead.facts or {},
        interest=lead.interest,
        confidence=lead.confidence,
        summary=lead.summary,
        open_concerns=lead.open_concerns or [],
        sent_items=lead.sent_items or [],
        call_attempts=lead.call_attempts,
        next_action_at=lead.next_action_at,
        last_whatsapp_inbound_at=lead.last_whatsapp_inbound_at,
        updated_at=lead.updated_at,
    )


@router.get("")
async def list_leads(
    status: str | None = Query(default=None),
    # is_lead filter: omit → all rows; true → real leads (default CRM screen);
    # false → raw data (the "Raw Data" view).
    is_lead: bool | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> list[LeadView]:
    leads = await get_store().list_leads(
        limit=limit, offset=offset, status=status, is_lead=is_lead
    )
    return [_to_view(x) for x in leads]


@router.post("")
async def create_lead(body: CreateLeadRequest) -> LeadView:
    store = get_store()
    # Upsert by phone so re-posting the same person doesn't duplicate.
    existing = await store.find_lead_by_phone(body.phone_e164) if body.phone_e164 else None
    if existing is not None:
        raise HTTPException(409, f"lead with phone already exists: {existing.lead_id}")
    lead = Lead(
        lead_id=body.lead_id or f"ld-{uuid.uuid4().hex[:8]}",
        full_name=body.full_name,
        email=body.email,
        phone_e164=(body.phone_e164 or "").strip().replace(" ", "").replace("-", ""),
        source=body.source,
        language_preference=body.language_preference,
        course_interest=body.course_interest,
        intake_year=body.intake_year,
        city=body.city,
        parent_name=body.parent_name,
        parent_phone_e164=body.parent_phone_e164,
        consent_call=body.consent_call,
        consent_whatsapp=body.consent_whatsapp,
        status=body.status,
        funnel_stage=body.funnel_stage,
        is_lead=body.is_lead,  # recomputed from funnel_stage in create_lead
    )
    created = await store.create_lead(lead)
    log.info("lead created", lead_id=created.lead_id)
    return _to_view(created)


# ---------------------------------------------------------------------------
# Bulk upload — .xlsx of leads → parse → insert (dialer picks them up)
# ---------------------------------------------------------------------------
# Map flexible spreadsheet headers (lowercased, spaces/underscores stripped) to
# Lead fields. Every field is OPTIONAL except the primary key, which we always
# auto-generate — a row with only a phone (or even fewer columns) still inserts;
# model defaults fill the NOT-NULL columns (full_name="Unknown", source, etc.).
_HEADER_ALIASES: dict[str, str] = {
    "name": "full_name", "fullname": "full_name", "full_name": "full_name",
    "candidate": "full_name", "student": "full_name", "studentname": "full_name",
    "phone": "phone_e164", "mobile": "phone_e164", "phonenumber": "phone_e164",
    "mobilenumber": "phone_e164", "contact": "phone_e164", "phonee164": "phone_e164",
    "whatsapp": "phone_e164",
    "email": "email", "emailid": "email", "mail": "email",
    "source": "source", "leadsource": "source",
    "language": "language_preference", "lang": "language_preference",
    "languagepreference": "language_preference", "preferredlanguage": "language_preference",
    "course": "course_interest", "courseinterest": "course_interest",
    "program": "course_interest", "branch": "course_interest", "interest": "course_interest",
    "intake": "intake_year", "intakeyear": "intake_year", "year": "intake_year",
    "city": "city", "location": "city", "town": "city",
    "parent": "parent_name", "parentname": "parent_name", "guardian": "parent_name",
    "parentphone": "parent_phone_e164", "parentmobile": "parent_phone_e164",
    "parentphonee164": "parent_phone_e164",
    # Admissions LIFECYCLE: the funnel stage column + the raw-data flag.
    "funnelstage": "funnel_stage", "funnelstatus": "funnel_stage",
    "stage": "funnel_stage", "status": "funnel_stage", "statusstage": "funnel_stage",
    "leadstage": "funnel_stage", "leadstatus": "funnel_stage",
    "islead": "is_lead", "lead": "is_lead", "raw": "is_lead", "rawdata": "is_lead",
}
_INT_FIELDS = {"intake_year"}

# Valid funnel_stage strings the sheet may carry. An unrecognised value is
# ignored (the row falls back to the default stage) rather than failing the row —
# so a typo never blocks an import.
_KNOWN_FUNNEL_STAGES = set(FunnelStage.ALL)

_TRUE_WORDS = {"true", "yes", "y", "1", "lead", "real"}
_FALSE_WORDS = {"false", "no", "n", "0", "raw", "rawdata", "raw_data"}


def _parse_bool(v: object) -> bool | None:
    """Parse a spreadsheet cell into a bool. None if blank/unrecognised (caller
    then uses the default). Handles 'yes'/'no', 'true'/'false', 1/0, and the
    domain words 'raw' (→ False) / 'lead' (→ True)."""
    if isinstance(v, bool):
        return v
    s = _cell(v)
    if s is None:
        return None
    s = s.strip().lower()
    if s in _TRUE_WORDS:
        return True
    if s in _FALSE_WORDS:
        return False
    return None


def _norm_header(h: object) -> str:
    return "".join(ch for ch in str(h or "").strip().lower() if ch.isalnum())


def _clean_phone(v: object) -> str:
    s = str(v or "").strip()
    # Spreadsheets often store phones as floats ("9.19e9") or with separators.
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")


def _cell(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


class UploadLeadsResult(BaseModel):
    inserted: int = 0
    duplicates: int = 0          # already existed (matched by phone) → skipped
    errors: int = 0
    rows: int = 0
    error_details: list[str] = []
    inserted_ids: list[str] = []


@router.post("/upload", response_model=UploadLeadsResult)
async def upload_leads(file: UploadFile = File(...)) -> UploadLeadsResult:
    """Bulk-insert leads from an Excel (.xlsx) file.

    Header row maps flexibly to lead fields (see _HEADER_ALIASES); unknown
    columns are ignored. Every field is optional — only the primary key is
    required and is auto-generated. Rows are upserted by phone: an existing
    phone is skipped (counted as a duplicate) rather than duplicated.

    Two optional admissions columns control the lifecycle:
      • Status / Stage  — a starting status (e.g. 'raw', 'cold', 'application_
                          started'). Unrecognised values are ignored.
      • Is Lead / Raw   — true/false (or 'lead'/'raw', yes/no, 1/0). False marks
                          a RAW-DATA row: inbound-only, never cold-called.
    Cross-defaults: status 'raw' ⇒ is_lead False; is_lead False ⇒ status 'raw';
    otherwise a real lead at status 'new' (so the dialer picks it up).
    """
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        raise HTTPException(400, "Please upload an Excel .xlsx file.")

    data = await file.read()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"Could not read the Excel file: {e}") from e

    # First non-empty row = headers.
    headers: list[str] | None = None
    for raw in rows_iter:
        if raw and any(c is not None and str(c).strip() for c in raw):
            headers = [_norm_header(c) for c in raw]
            break
    if not headers:
        raise HTTPException(400, "The sheet appears to be empty.")
    # Column index → lead field name (only recognised columns).
    col_field = {i: _HEADER_ALIASES[h] for i, h in enumerate(headers) if h in _HEADER_ALIASES}
    if not col_field:
        raise HTTPException(
            400,
            "No recognisable columns found. Include at least a name or phone "
            "column (e.g. 'Name', 'Phone'/'Mobile').",
        )

    store = get_store()
    out = UploadLeadsResult()

    for rownum, raw in enumerate(rows_iter, start=2):
        if not raw or not any(c is not None and str(c).strip() for c in raw):
            continue  # skip blank rows
        out.rows += 1
        fields: dict[str, object] = {}
        is_raw_flag: bool | None = None      # from an is_lead/raw column
        funnel_val: str | None = None        # from a stage/status column
        for i, field_name in col_field.items():
            if i >= len(raw):
                continue
            val = raw[i]
            if field_name == "phone_e164":
                val = _clean_phone(val)
            elif field_name in _INT_FIELDS:
                cv = _cell(val)
                try:
                    val = int(float(cv)) if cv else None
                except ValueError:
                    val = None
            elif field_name == "is_lead":
                # Column says whether it's a lead. False / "raw" → raw data.
                b = _parse_bool(val)
                if b is not None:
                    is_raw_flag = not b
                continue  # handled separately
            elif field_name == "funnel_stage":
                cv = (_cell(val) or "").strip().lower().replace(" ", "_")
                funnel_val = cv if cv in _KNOWN_FUNNEL_STAGES else None
                continue  # handled separately
            else:
                val = _cell(val)
            if val not in (None, ""):
                fields[field_name] = val

        # Resolve the funnel stage with sensible cross-defaults:
        #   - explicit stage "raw"          → raw data
        #   - explicit is_lead False / raw  → stage "raw"
        #   - otherwise                     → a real lead at stage "lead"
        if funnel_val is None:
            funnel_val = FunnelStage.RAW if is_raw_flag else FunnelStage.LEAD
        elif funnel_val != FunnelStage.RAW and is_raw_flag:
            # Conflicting signals (a non-raw stage but raw flagged) → trust the
            # raw flag, the more explicit "this isn't a lead" intent.
            funnel_val = FunnelStage.RAW

        phone = str(fields.get("phone_e164", "") or "")
        try:
            # Upsert by phone — don't duplicate an existing lead.
            if phone:
                existing = await store.find_lead_by_phone(phone)
                if existing is not None:
                    out.duplicates += 1
                    continue
            lead = Lead(
                lead_id=f"ld-{uuid.uuid4().hex[:8]}",
                source=str(fields.pop("source", "upload") or "upload"),
                status=LeadStatus.NEW,         # operational state — dialer-eligible
                funnel_stage=funnel_val,       # lifecycle from sheet, or default
                is_lead=is_lead_for_stage(funnel_val),  # derived (also re-set in create_lead)
                # An uploaded sheet IS the consented outbound call list — the
                # model defaults are false, which would make every uploaded
                # lead invisible to the dialer (and skip WhatsApp follow-ups).
                consent_call=True,
                consent_whatsapp=True,
                **fields,                      # all other fields optional
            )
            await store.create_lead(lead)
            out.inserted += 1
            out.inserted_ids.append(lead.lead_id)
        except Exception as e:  # noqa: BLE001
            out.errors += 1
            if len(out.error_details) < 25:
                out.error_details.append(f"row {rownum}: {e}")

    log.info(
        "leads bulk upload",
        file=file.filename, rows=out.rows,
        inserted=out.inserted, duplicates=out.duplicates, errors=out.errors,
    )
    return out


class ResetResult(BaseModel):
    ok: bool = True
    cleared: dict[str, int] = {}     # rows removed per table: {leads, sessions, tasks}


@router.post("/reset", response_model=ResetResult)
async def reset_leads() -> ResetResult:
    """DESTRUCTIVE: clear ALL rows from leads, sessions and tasks (the tables
    themselves stay). Used by the CRM 'Reset' button to wipe the operational
    data for a fresh start. Equivalent to:
        TRUNCATE TABLE tasks, sessions, leads RESTART IDENTITY CASCADE
    """
    cleared = await get_store().reset_all()
    log.warning("CRM RESET — operational tables truncated", cleared=cleared)
    return ResetResult(ok=True, cleared=cleared)


class RecordDeliveryRequest(BaseModel):
    item: str
    channel: str = "whatsapp"


class ScheduleFollowupRequest(BaseModel):
    in_minutes: int = 1440  # default: next day


@router.post("/{lead_id}/schedule-followup")
async def schedule_followup(lead_id: str, body: ScheduleFollowupRequest) -> dict:
    """Put the lead back in the dial queue after `in_minutes` (status FOLLOWUP +
    next_action_at). Used by AegisBackend when an outbound call goes unanswered
    (busy / rejected / no-answer) so the candidate is retried the next day.
    Terminal leads are ignored by the store."""
    lead = await get_store().get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    await get_store().schedule_followup(
        lead_id=lead_id, in_minutes=max(1, int(body.in_minutes))
    )
    log.info("followup scheduled via api", lead_id=lead_id, in_minutes=body.in_minutes)
    return {"ok": True}


class AdvanceStageRequest(BaseModel):
    # One of the FunnelStage values (lead / application_started / fees_pending /
    # application_submitted / raw). is_lead is derived from it.
    funnel_stage: str
    # Whether to fire the post-payment next-steps WhatsApp (entrance-exam slots
    # vs token amount). None → auto (fires for fees_pending + application_submitted).
    send_next_steps: bool | None = None


@router.post("/{lead_id}/advance-stage")
async def advance_stage(lead_id: str, body: AdvanceStageRequest) -> LeadView:
    """SINGLE entry point to move a lead through the admissions LIFECYCLE
    (funnel_stage, Module F). Any source — a portal/payment webhook, a counsellor
    clicking a stage in the CRM, or the AI inferring it from a call — POSTs here.
    When the payment is confirmed (send_next_steps True, or auto for the
    payment-done stages) the course-appropriate next-steps message is queued:
    entrance-exam phases + slot booking, or the token-amount booking link."""
    lead = await get_store().get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    if body.funnel_stage.strip().lower() not in FunnelStage.ALL:
        raise HTTPException(
            422,
            f"unknown funnel_stage={body.funnel_stage!r}; "
            f"expected one of {sorted(FunnelStage.ALL)}",
        )
    updated = await get_store().advance_stage(
        lead_id=lead_id, funnel_stage=body.funnel_stage, send_next_steps=body.send_next_steps
    )
    if updated is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    log.info("lead stage advanced via api", lead_id=lead_id, funnel_stage=body.funnel_stage)
    return _to_view(updated)


@router.post("/{lead_id}/deliveries")
async def record_delivery(lead_id: str, body: RecordDeliveryRequest) -> dict:
    """Record that something was delivered to the candidate (e.g. a document
    sent live on WhatsApp by the brain's send_document tool). Appended to
    `sent_items` so the next conversation follows up instead of re-offering —
    same loop the post-call action worker uses."""
    lead = await get_store().get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    await get_store().record_delivery(lead_id=lead_id, item=body.item, channel=body.channel)
    return {"ok": True}


@router.get("/by-phone/{phone}")
async def get_lead_by_phone(phone: str) -> LeadView:
    """Resolve a lead by phone number — used by AegisBackend on an INBOUND call
    to identify the caller (and their lead_id) before answering, so the opener
    greets by name and memory hydrates. `phone` may be digits-only or E.164; the
    store matches either form. 404 when the caller is unknown."""
    lead = await get_store().find_lead_by_phone(phone)
    if lead is None:
        raise HTTPException(404, f"no lead with phone={phone!r}")
    return _to_view(lead)


@router.get("/{lead_id}/sessions")
async def get_lead_sessions(lead_id: str) -> list[dict]:
    """All sessions for a candidate (most-recent first), for the detail page.
    Each carries channel/direction, status + end_reason, turn count, the
    per-session analysis (sentiment/summary/actions), and timestamps. 404 if the
    lead itself is unknown."""
    store = get_store()
    lead = await store.get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    sessions = await store.list_sessions_for_lead(lead_id)
    return [
        {
            "session_id": s.session_id,
            "channel": s.channel,
            "direction": s.direction,
            "status": s.status,
            "end_reason": s.end_reason,
            "analyzed": s.analyzed,
            "turns": len(s.transcript or []),
            "transcript": s.transcript or [],
            "analysis": s.analysis,
            "started_at": s.started_at,
            "ended_at": s.ended_at,
        }
        for s in sessions
    ]


@router.get("/{lead_id}/memory")
async def get_memory(lead_id: str) -> MemoryBundle:
    """The agent-memory read-model. Cheap, read once at session start."""
    lead = await get_store().get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    return MemoryBundle(
        lead_id=lead.lead_id,
        full_name=lead.full_name,
        language_preference=lead.language_preference,
        status=lead.status,
        funnel_stage=lead.funnel_stage,
        lead_priority=lead.lead_priority,
        interest=lead.interest,
        confidence=lead.confidence,
        facts=lead.facts or {},
        summary=lead.summary,
        open_concerns=lead.open_concerns or [],
        sent_items=lead.sent_items or [],
    )


@router.get("/{lead_id}")
async def get_lead(lead_id: str) -> LeadView:
    lead = await get_store().get_lead(lead_id)
    if lead is None:
        raise HTTPException(404, f"unknown lead_id={lead_id!r}")
    return _to_view(lead)
